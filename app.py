# cd C:\Users\a2907\Desktop\pyApp\DeepLearning\Render\app1
# streamlit run app.py
#
import pandas as pd
import streamlit as st
import plotly.express as px
from PIL import Image
import openpyxl
from openpyxl import load_workbook

# extract Excel Data Create Panda Dataframe
# df=extractExcelDataCreatePandaData(excel_file,sheet_name,start_col,end_col,header_rows)
def extractExcelDataCreatePandaData(excel_file,sheet_name,start_col,end_col,header_rows):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]
    # get header
    header_title = [cell.value for cell in ws[header_rows]]
    header_Name = header_title[start_col - 1:end_col]
    # Create an empty list to hold the data
    data = []
    # Iterate through the rows in the selected range
    for row in ws.iter_rows(min_col=start_col, max_col=end_col, min_row=header_rows + 1, values_only=True):
        data.append(row)
    # Create a Pandas DataFrame from the data
    df = pd.DataFrame(data, columns=header_Name)
    # data1 = list(map(list, zip(*data))), transpose list
    return df
#
st.set_page_config(page_title='Survey Results')
st.header('Survey Results 2021')
st.subheader('Was the tutorial helpful?')
#
# extract Excel Data Create Panda Dataframe
excel_file = 'Survey_Results.xlsx'
sheet_name = 'DATA'
start_col=2
end_col=4
header_rows = 4
# header_Name = ['Department', 'Age', 'Rating']
df=extractExcelDataCreatePandaData(excel_file,sheet_name,start_col,end_col,header_rows)
#
start_col=6
end_col=7
# header_Name = ['Departments', 'Participants']
df_participants=extractExcelDataCreatePandaData(excel_file,sheet_name,start_col,end_col,header_rows)
df_participants.dropna(inplace=True)

# --- STREAMLIT SELECTION
department = df['Department'].unique().tolist()
ages = df['Age'].unique().tolist()

age_selection = st.slider('Age:',
                        min_value= min(ages),
                        max_value= max(ages),
                        value=(min(ages),max(ages)))

department_selection = st.multiselect('Department:',
                                    department,
                                    default=department)

# --- FILTER DATAFRAME BASED ON SELECTION
mask = (df['Age'].between(*age_selection)) & (df['Department'].isin(department_selection))
number_of_result = df[mask].shape[0]
st.markdown(f'*Available Results: {number_of_result}*')

# --- GROUP DATAFRAME AFTER SELECTION
df_grouped = df[mask].groupby(by=['Rating']).count()[['Age']]
df_grouped = df_grouped.rename(columns={'Age': 'Votes'})
df_grouped = df_grouped.reset_index()

# --- PLOT BAR CHART
bar_chart = px.bar(df_grouped,
                   x='Rating',
                   y='Votes',
                   text='Votes',
                   color_discrete_sequence = ['#F63366']*len(df_grouped),
                   template= 'plotly_white')
st.plotly_chart(bar_chart)

# --- DISPLAY IMAGE & DATAFRAME
col1, col2 = st.columns(2)
image = Image.open('images/survey.jpg')
col1.image(image,
        caption='Designed by slidesgo / Freepik',
        use_column_width=True)
col2.dataframe(df[mask])

# --- PLOT PIE CHART
pie_chart = px.pie(df_participants,
                title='Total No. of Participants',
                values='Participants',
                names='Departments')

st.plotly_chart(pie_chart)
